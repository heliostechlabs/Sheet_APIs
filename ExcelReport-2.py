import openpyxl

def read_data_from_source_file(source_file, sheet_name, column_index):
    
    # the code assumes that row 1 contains the column header and the data begins from row 2
    workbook = openpyxl.load_workbook(source_file)
    sheet = workbook[sheet_name]
    column_data = []
    data_dict = {}
    sales_dict = {}
    bills_dict = {}
    accounts_dict = {}
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=column_index, max_col=8, values_only=True):
        
        if (row[2] == "Expense" or row[2] == "Bills"):
        
            accounts_dict.update({row[4]: row[1]})
            
        if row[1] in data_dict:
            new_val = row[0]
            updated_val = new_val + data_dict.get(row[1])
            data_dict.update({row[1]: updated_val})
            
            if row[2] == "Bills":
                new_amount = row[0]
                updated_val = new_amount + bills_dict.get(row[1], 0)
                bills_dict.update({row[1]: updated_val})

        else:
            data_dict.update({row[1]: row[0]})
                
        column_data.append(row[1])
    
    columns = dict.fromkeys(column_data)
    
    list_accounts_dict = list(accounts_dict.keys())
    
    add_columns_at_location('C:\\Users\\brady\\Desktop\\Book1.xlsx', 'Sheet1', 4, 2, columns, bills_dict, data_dict, list_accounts_dict)
        
    return column_data


def add_columns_at_location(file_path, sheet_name, column_index, num_columns, column_names, bills_dict, data_dict, accounts_dict):
 
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.insert_cols(column_index, amount=num_columns)
    
    for i, column_name in enumerate(column_names):
        sheet.cell(row=1, column=column_index + i, value=column_name)
        # sheet.cell(row=6, column=column_index + i, value=bills_dict.get(column_name))
        # sheet.cell(row=3, column=column_index + i, value=data_dict.get(column_name) - bills_dict.get(column_name, 0))

    for i, account_name in enumerate(accounts_dict):
        sheet.cell(row=2+i, column=2, value=account_name)
        
        
    # for i, column_name in enumerate(column_names):
    #     sheet.cell(row=3, column=column_index + i, value=column_name)
        
    workbook.save(file_path)
    

sheet_name = 'Sheet1'
source_sheet_name = 'Sheet1'
source_file = "C:\\Users\\brady\\Desktop\\ZohoData.xlsx"
column_names = read_data_from_source_file(source_file, source_sheet_name, 2)

# final_col_names = []

# for column in column_names:
#     if column != None:
#         final_col_names.append(column)

try1 = list(dict.fromkeys(column_names))

file_path = 'C:\\Users\\brady\\Desktop\\Book1.xlsx'
column_index =  3 # 1 indexing is followed
num_columns = 2   
# column_names = ['NewColumn3', 'NewColumn4']

# col_names = list(data_dic.keys())
# add_columns_at_location(file_path, sheet_name, column_index, num_columns, try1)
